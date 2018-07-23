#engagement while immersed in a virtual world

##look at facial analysis engagement, suprise, excitement
##look at pulse
##command + space -- 'managed software center' -- install

#what do we need to do
#1- get top n gsr peaks from each participant
#DONEDONE
#2- do analysis on data in a 20 second window around the top n peaks
#3- transplant this analysis into the excel sheet

#building blocks for first goal

#make box and whisker plot
#get gsr baseline

#make audio-gsr graph similar to gsr peak analysis.
    #-if a speaking interval has a gsr peak, plot 1 in that inverval.  for each participant
    #-compare vr vs. monitor conditions
    #-try get t-tests for these


# get first and second half gsr values across groups
import openpyxl
import pandas as pd
from openpyxl import load_workbook
import numpy as np
from matplotlib.ticker import MaxNLocator
import matplotlib.mlab as mlab
import matplotlib.pyplot as plt

from pylab import *
import pylab
from scipy.io import wavfile
import os
import more_itertools as mit

#importing the workbook.  Change this filepath for where the data is on your machine.
wb = openpyxl.load_workbook('C:\\Users\\Jake From State Farm\\Desktop\\Peaks pr respondent.xlsx')
data_folder = 'C:/Users/Jake From State Farm/AppData/Local/Programs/Python/Python36-32/Sensor Data'
sheet = wb['Peaks pr respondent']

performance_length = 336000

####################################AUDIO########################################
def find_ranges(iterable):
    """Yield range of consecutive numbers."""
    for group in mit.consecutive_groups(iterable):
        group = list(group)
        if len(group) == 1:
            yield int(1000 * group[0]), int(1000 + 1000 * group[0])
        else:
            yield int(1000* group[0]), int(1000 + 1000*group[-1])

def signalAboveThresh(audioSnip, threshold):
    continuousSignal = True
    #list comprehension from current position to 50 places out
    for amp in audioSnip:
        if amp < threshold:
                #if our amplitude drops below threshold, our signal is not continuous
                continuousSignal = False
                break
    #only append to list if this is true!!!!
    return continuousSignal

def singleParRange(num, doTransformOutput):
        soundTime = []
        counter = 0
        maxCount = 50
        threshold = 1000


    #GETTING SOUND DATA
         #tori path:
        filePathwayRead = os.path.expanduser("~/Desktop/SilenceTracker/TrimmedAudioFiles/Participant"+str(num)+"Trim.wav")
        #Jake path, just comment this out for the code to work on your end.  
        filePathwayRead = os.path.expanduser("C:/Users/Jake From State Farm\Desktop/TrimmedAudioFiles/Participant"+str(num)+"Trim.wav")
        sampFreq, theSoundFile = wavfile.read(filePathwayRead)
        print("Opening participant " + str(num) + " Audio file. . . " + filePathwayRead)

     #ANALYZING DATA
        print("Getting participant " + str(num) + "'s Spoken Times. . . ")
        for index, amp in enumerate(theSoundFile):
            #checking every 10th ms
            if index % 10 == 0:
                if abs(amp) > threshold:
                    theSecond = round(index/100000, 0)
                    
                    #checking if our second is already in the list.  No need to check if
                    #we have already analyzed this second and found audio signals
                    if theSecond in soundTime:
                        continue
                    
                    #this function looks awful but its only checking if our signal
                    #holds for more than 50 milliseconds (this paramater under maxCount)
                    #if yes then we append this second to the list
                    elif signalAboveThresh(theSoundFile[index: index + maxCount], threshold) == True:
                        soundTime.append(theSecond)
                            
        #Changing the list format
        if doTransformOutput
        return list(find_ranges(soundTime))

    
#obtaining file
def getParRanges():
    
    participantSoundTime = {}

    for num in range(1, 4):
        if num == 11 or num == 13 or num == 14 or num == 16 or num == 21 or num == 22 or num == 26 or num == 27:
            continue
        else:
            #adding completed list of times to a dictionary where participant # is key      
            participantSoundTime[str(num)] = singleParRange(num)
        
    return participantSoundTime

###########################################################################
#################GETTING HEARTRATE DERIVATIVE#############################
###########################################################################

#yvals is a list or tuple of 2 values
def getSlope(yvals, xdist):
    return (yvals[1] - yvals[0]) / xdist

#many_yvals is a list of arbitrary size
def getDerivative(many_yvals, xdist):
    #returns a list, may make it just write to file to save memory
    deriv = []
    previous = 0
    for v in many_yvals:
        if v != None and v > 0:
            deriv.append(getSlope((previous, v), xdist))
            previous = v
        else:
            deriv.append(0)
            previous = 0

    deriv_dict = {}
    for i, d in enumerate(deriv):
        if d != 0:
            #making dictionary of structure {time(ms) : derivitive}
            #this shows us when the derivitive changes and by how much
            deriv_dict[i] = d
    return deriv_dict

def getSingleDer(file, xdist):
        data = []
        path = data_folder + '/' +  file
        wrkbk = load_workbook(filename=path)
        sheet = wrkbk.sheetnames[0]
        sheet = wrkbk[sheet]

        prev = 0
        for cell in sheet['FJ']:
            try:
                cell.value += 1
            except TypeError:
                continue
            data.append(cell.value)
        print(data[0])
        return getDerivative(data, xdist)

def getAllDer(xdist):
        #this may complain about double \\ in data_folder dir name
        for file in os.listdir(data_folder):
                getSingleDer(file, 1)
                
if 0:
    min_to_ms = 1.66 * (10**-5)
    print(getSingleDer('001_Participant 1.xlsx', 1))

###########################################################################
########################FINDING GSR PEAKS##################################
###########################################################################
def makeBins(n_bins):
    bin_length = performance_length / n_bins
    bins = []
    cur_val = 0
    for i in range(n_bins):
        bins.append((cur_val, cur_val + bin_length))
        cur_val += bin_length
    return bins

def getCondition(participant):
    #0 for monitor, 1 for VR
     for cell in sheet['E']:
        if cell.value == ('Participant ' + str(participant)):
            row = cell.row
            if sheet['H' + str(row)].value == 'Virtual Reality':
                return 1
            else:
                return 0


def isPeak(participant, time_range, activation_threshold,single_or_total):
    #time_range is a 2-tuple, like how makeBins outputs
    val = 0
    for cell in sheet['E']:
        if cell.value == ('Participant ' + str(participant)):
            row = cell.row
            #checking if peakMs is within time_range
            if sheet['K' + str(row)].value > time_range[0] and sheet['k' + str(row)].value < time_range[1]:
                if sheet['L' + str(row)].value > activation_threshold:
    #data is returned as (peak signal, study group) tuple
                    if single_or_total == 0:
                        return 1
                    else:
                        val += 1
    if single_or_total == 0:
        return 0
    else:
        return val

def histogramGSRData(condition, activation_threshold, n_bins, single_or_total):
    #loop through participants
    bins = makeBins(n_bins)
    data = [0 for x in range(len(bins))]
    for participant in range(1, 33):
        #check if we are in the right condition
        if getCondition(participant) == condition:
            for j in range(len(bins)):
                data[j] += isPeak(participant, bins[j], activation_threshold, single_or_total)
    return data

def histogramAudioData(condition, 

if 0:
    ##PARAMETERS##
    single_or_total = 0
    timestep = 10
    threshold = 0.25
    ################

    my_bins = int(336/timestep)
    x1 = histogramGSRData(1, threshold, my_bins, single_or_total)
    x2 = histogramGSRData(0, threshold, my_bins, single_or_total)
    bar_width = .3* timestep

    bins = makeBins(my_bins)
    y1 = [int(i[0] / 1000) for i in bins]
    y2 = [i + bar_width for i in y1]

    ##y1 = [i for i in range(my_bins)]
    ##y2 = [i + bar_width for i in y1]


    error_config = {'ecolor': '0.3'}

    fig, ax = plt.subplots()
    opacity = 0.4
    rects1 = ax.bar(y1, x1, bar_width, error_kw=error_config, label='VR')
    rects2 = ax.bar(y2 , x2, bar_width, error_kw=error_config, label='Monitor')

    ax.set_xlabel('Time (s)')
    ax.set_ylabel('Number of Participants with peaks above' + str(threshold))
    ax.set_title('GSR Peaks per ' + str(timestep) + ' Second Bin')
    ax.yaxis.set_major_locator(MaxNLocator(integer=True))
    ax.legend()

    fig.tight_layout()
    plt.show()



###print(getCondition(1))
###print(makeBins(10))
##bins = int(336/10)
##thresholds = [.25, .5, .75]
##for i in range(len(thresholds)):
##    print('monitor condition with threshold of ' + str(thresholds[i]) + ' ' + str(histogramData(0, thresholds[i], bins)))
##    print('VR condition with threshold of ' + str(thresholds[i]) +  ' ' + str(histogramData(1, thresholds[i], bins)))

###########################GSR MEAN SCORES###########################################
##using pandas library here, trying out different libraries to see which ones i prefer
def complimentIntervals(timeRange, listOfIntervals):
    compliment = []
    lastval = 0
    for interval in listOfIntervals:
        if interval[0] != 0:
            compliment.append((lastval, interval[0]))
        lastval = interval[1]
    return compliment

    

def getSingleGSRAvg(times, df):
        initTime = times[0]
        endTime = times[1]
        df = df.iloc[initTime:endTime]
        print('times are: ' + str(times))
        print('Calculating GSR Mean in your interval...')
        try:
            vals = df.mean()
            print(vals[0])
            return vals[0]
        except:
            return -1

def getAllAvg(times, condition):
        #this may complain about double \\ in data_folder dir name
        arr = []
        for i, file in enumerate(os.listdir(data_folder)):
                if getCondition(i) == condition:
                    arr.append(getSingleGSRAvg(file, times[0], times[1]))
        return arr
    
def getGSR(times, df):
    initTime = times[0]
    endTime = times[1]
    x = df.iloc[initTime:endTime]
    return x

def boxPlot(participant, ranges):
    #getting file
    file = os.listdir(data_folder)[participant]
    path = data_folder + '/' +  file
    #print('Opening GSR File.  This could take a few moments.')
    df = pd.read_excel(path, header = None, usecols = [167])
    
    data = []
    column = []
    for tupl in ranges:
        if getGSR(tupl, df).empty:
            continue
        else:
            x = (tupl[0]/1000, tupl[1]/1000)
            column.append(str(x))
            data.append(getGSR(tupl, df))

    df = pd.concat(data, axis = 1)
    df.columns = column
    df.plot.box()
    plt.show()
    
def GSRAudioInterest(participant, ranges):
    #getting file
    averages = []
    file = os.listdir(data_folder)[participant]
    path = data_folder + '/' +  file
    print('Opening GSR File.  This could take a few moments.')
    df = pd.read_excel(path, header = None, usecols = [167])
    #print(df)
    for tupl in ranges:
        #print('your tuple is: ' + str(tupl))
        averages.append(getSingleGSRAvg(tupl, df))
        
    return averages

participant = 3
ranges = singleParRange(participant)
rangesCompliment = complimentIntervals(performance_length, ranges)
boxPlot(participant, ranges)
#boxPlot(participant, rangesCompliment)




